import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from datetime import datetime
import re


###########################################################
def extract_data_from_pdf(pdf_path, text_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            ##
            chars = page.chars # 提取每個字元的座標和內容

            extracted_text = ""
            prev_char = None

            for char in chars:
                if prev_char:
                    # 根據 x 座標間距判斷是否插入空格，這裡使用一個簡單的閾值 1.0
                    if char['x0'] - prev_char['x1'] > 2.0:
                        extracted_text += ' '

                    # 根據 y 座標變化判斷是否需要換行，這裡使用一個簡單的閾值 5.0
                    if abs(char['top'] - prev_char['top']) > 5.0:
                        extracted_text += '\n'

                extracted_text += char['text']
                prev_char = char

            print(extracted_text)

            text += extracted_text

            # text += page.extract_text()
    #with open(text_path, 'w') as f:
    with open(text_path, 'w', encoding='utf-8') as f:
        f.write(text)
###############################################################

# def extract_data_from_pdf(pdf_path, text_path):
#     with pdfplumber.open(pdf_path) as pdf:
#         text = ""
#         for page in pdf.pages:
#             text += page.extract_text()
#     #with open(text_path, 'w') as f:
#     with open(text_path, 'w', encoding='utf-8') as f:
#         f.write(text)


def clean_description(description):
    while "Page" in description and "Nett Price Value" in description:
        start_idx = description.find("Page")-1
        ## end_idx = description.find("NettPriceValue") + len("NettPriceValue")
        end_idx = description.find("Nett Price Value") + len("Nett Price Value")+1  # 新版本Nett Price 空格有分辨出來
        description = description[:start_idx] + description[end_idx:]
    if "******" in description:
        description = description.split("******")[0]
        # description = description.replace('\n', '\\n') # 修正最後一筆資料不正確換行符號問題
    return description


def parse_text_to_data(text_path):
    data = []
    order_number = ""
    order_date = ""
    item_counter = 1

    with open(text_path, 'r') as f:
        lines = f.readlines()

    # 寻找订单号和日期
    for line in lines:
        if "ORDER NUMBER" in line:
            match = re.search(r"ORDER NUMBER\s*:\s*(\d+)", line)
            if match:
                order_number = match.group(1)
        
        if "TO: DATE :" in line:
            match = re.search(r"TO: DATE\s*:\s*(\d{2}/\d{2}/\d{2})", line)
            if match:
                # 转换日期格式为 YY/MM/DD
                date_str = match.group(1)
                date_obj = datetime.strptime(date_str, '%d/%m/%y')
                order_date = date_obj.strftime('%y/%m/%d')  # 修改这里的格式
                break

    current_product = None
    for line in lines:
        if line.startswith('FL') and not line.startswith('FLANG'):
            if current_product:
                current_product["Description"] = clean_description(current_product["Description"])
                data.append(current_product)
            current_product = {
                "Product No": "",
                "Quantity": "",
                "Units": "",
                "Description": "",
                "QA Req": "",
                "Date Req": "",
                "Nett Price": "",
                "Value": "",
                "PO#": order_number,
                "Date": order_date,
                "Item": str(item_counter)
            }
            item_counter += 1
            parts = line.split()
            current_product["Product No"] = parts[0]
            current_product["Quantity"] = int(float(parts[1]))
            current_product["Units"] = parts[2]
            current_product["Description"] = ' '.join(parts[3:-3]) + '>'

            # 修改日期处理部分
            date_str = parts[-3]
            date_obj = datetime.strptime(date_str, '%d/%m/%y')
            formatted_date = date_obj.strftime('%Y/%m/%d')
            current_product["Date Req"] = formatted_date

            current_product["Nett Price"] = round(float(parts[-2]), 2)
            # 修改 Value 的计算方式
            nett_price = round(float(parts[-2]), 2)
            quantity = int(float(parts[1]))
            current_product["Value"] = round(nett_price * quantity, 2)  # Value = Nett Price * Quantity

        elif line.strip().startswith('TOTAL'):
            continue
        elif current_product:
            current_product["Description"] += " " + '\n' + line.strip()

    if current_product:
        current_product["Description"] = clean_description(current_product["Description"])
        data.append(current_product)

    return data



def write_to_excel(data, excel_path):
    # 定义列的顺序
    columns_order = [
        "PO#",
        "Date",
        "Item",
        "Product No",
        "",  # 保留第一个空格栏位
        "Description",
        "QA Req",
        "Nett Price",
        "Quantity",
        "Date Req",
        "Value"
    ]

    # 创建 DataFrame 并添加空格栏位
    df = pd.DataFrame(data)
    df[""] = ""  # 只保留第一个空格栏位
    
    # 重新排序列
    df = df[columns_order]

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # 设置 PO# 列（第1列）为文本格式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = '@'

    # 设置 Date 列（第2列）为日期格式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = 'm/d/yy'

    # 设置 Date Req 列（第10列）为日期格式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=10):
        for cell in row:
            cell.number_format = 'm/d/yy'

    # 设置 Item 列和 Quantity 列的对齐方式
    for row in ws.iter_rows(min_row=2):
        # Item 列现在是第3列
        item_cell = row[2]
        item_cell.alignment = Alignment(horizontal='center', vertical='bottom')
        
        # Quantity 列现在是第8列（原来是第9列）
        quantity_cell = row[8]
        quantity_cell.alignment = Alignment(horizontal='center', vertical='bottom')

    # 设置 Description 列的格式（第6列，原来是第7列）
    description_column = 'F'  # 从G列改为F列
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.number_format = '@'
            num_lines = len(str(cell.value).split('\n'))
            ws.row_dimensions[cell.row].height = None

    # 设置 Description 列宽
    ws.column_dimensions[description_column].width = 50

    # 保存工作簿
    wb.save(excel_path)


def main():
    pdf_path = input("请输入PDF文件名：")
    if not pdf_path.endswith(".pdf"):
        pdf_path += ".pdf"

    excel_path = input("请输入输出Excel文件名：")
    if not excel_path:
        excel_path = pdf_path.replace(".pdf", ".xlsx")
    elif not excel_path.endswith(".xlsx"):
        excel_path += ".xlsx"

    text_path = "pdf2text.txt"

    extract_data_from_pdf(pdf_path, text_path)
    data = parse_text_to_data(text_path)
    write_to_excel(data, excel_path)
    print(f"数据已成功写入 {excel_path}")


if __name__ == "__main__":
    main()
